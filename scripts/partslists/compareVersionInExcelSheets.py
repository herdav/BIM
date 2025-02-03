# compareVersionInExcelSheets.py
# Created by David Herren
# Version 2025-02-03

import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from pathlib import Path
from openpyxl.styles import Font, Border

def select_file(title):
  root = tk.Tk()
  root.withdraw()
  file_path = filedialog.askopenfilename(title=title, filetypes=[("Excel files", "*.xlsx")])
  return file_path

def auto_adjust_column_width(worksheet, columns):
  # Automatically adjust column widths
  for col in columns:
    max_length = 0
    column = worksheet[col]
    for cell in column:
      if cell.coordinate in worksheet.merged_cells:
        continue
      try:
        if cell.value and len(str(cell.value)) > max_length:
          max_length = len(str(cell.value))
      except:
        pass
    adjusted_width = max_length + 2
    worksheet.column_dimensions[col].width = adjusted_width

def main():
  try:
    old_file = select_file("Select the old Excel file")
    new_file = select_file("Select the new Excel file")

    old_file_name = Path(old_file).name
    new_file_name = Path(new_file).name

    # Read Excel files with header in the second row (header=1)
    df_old = pd.read_excel(old_file, header=1)
    df_new = pd.read_excel(new_file, header=1)

    # Clean column names by stripping whitespace
    df_old.columns = df_old.columns.str.strip()
    df_new.columns = df_new.columns.str.strip()

    required_columns = ['Art.No.', 'Bem.-Remarques', 'Stk', 'Bezeich.-Désignation']
    for df in [df_old, df_new]:
      missing_columns = [col for col in required_columns if col not in df.columns]
      if missing_columns:
        raise ValueError(f"Missing columns in data: {missing_columns}")

    # Remove extra spaces and replace NaN/'nan' with empty strings for Art.No. and Bem.-Remarques
    df_old['Art.No.'] = df_old['Art.No.'].astype(str).str.strip().replace('nan', '')
    df_old['Bem.-Remarques'] = df_old['Bem.-Remarques'].astype(str).str.strip().replace('nan', '')
    df_new['Art.No.'] = df_new['Art.No.'].astype(str).str.strip().replace('nan', '')
    df_new['Bem.-Remarques'] = df_new['Bem.-Remarques'].astype(str).str.strip().replace('nan', '')

    # Create merge key:
    # If either Art.No. or Bem.-Remarques is non-empty, use "Art.No.|Bem.-Remarques" as key.
    # Otherwise (both are empty), use Bezeich.-Désignation as the key.
    def create_key(row):
      art = str(row['Art.No.']).strip() if pd.notnull(row['Art.No.']) else ''
      bem = str(row['Bem.-Remarques']).strip() if pd.notnull(row['Bem.-Remarques']) else ''
      designation = str(row['Bezeich.-Désignation']).strip() if pd.notnull(row['Bezeich.-Désignation']) else ''
      if art or bem:
        return art + '|' + bem
      else:
        return designation
    df_old['Key'] = df_old.apply(create_key, axis=1)
    df_new['Key'] = df_new.apply(create_key, axis=1)

    # Filter columns for merging
    df_old = df_old[['Key', 'Stk', 'Bezeich.-Désignation']]
    df_new = df_new[['Key', 'Stk', 'Bezeich.-Désignation']]

    # Merge dataframes on the created key
    df_merged = pd.merge(df_old, df_new, on='Key', how='outer', suffixes=('_old', '_new'), indicator=True)

    # Determine change status
    df_merged['Änderung'] = df_merged.apply(
      lambda row: 'Entfernt' if row['_merge'] == 'left_only'
                  else 'Hinzugefügt' if row['_merge'] == 'right_only'
                  else ('Geändert' if row['Stk_old'] != row['Stk_new'] else 'Keine Änderung'),
      axis=1
    )

    # Split the key to retrieve Art.No. and Bem.-Remarques if possible.
    def split_key(key):
      key = str(key)  # ensure key is a string
      if '|' in key:
        parts = key.split('|', 1)
        return pd.Series(parts)
      else:
        return pd.Series(["", ""])
    df_merged[['Art.No.', 'Bem.-Remarques']] = df_merged['Key'].apply(split_key)

    # Drop helper columns
    df_merged = df_merged.drop(columns=['Key', '_merge'])

    # Update Bezeich.-Désignation: for removed entries, use the old designation;
    # otherwise, use the new designation.
    df_merged['Bezeich.-Désignation'] = df_merged.apply(
      lambda row: row['Bezeich.-Désignation_old'] if row['Änderung'] == 'Entfernt' else row['Bezeich.-Désignation_new'],
      axis=1
    )

    # Sort data (order: Hinzugefügt -> Geändert -> Entfernt -> Keine Änderung)
    sort_order = {'Hinzugefügt': 0, 'Geändert': 1, 'Entfernt': 2, 'Keine Änderung': 3}
    df_merged = df_merged.sort_values(
      by=['Änderung', 'Art.No.', 'Bem.-Remarques'],
      key=lambda col: col.map(sort_order).fillna(col)
    )

    # Rename columns: Stk_old -> Stk.-alt, Stk_new -> Stk.-neu
    df_merged.rename(columns={'Stk_old': 'Stk.-alt', 'Stk_new': 'Stk.-neu'}, inplace=True)

    # Select desired columns
    df_merged = df_merged[['Änderung', 'Stk.-alt', 'Stk.-neu', 'Art.No.', 'Bezeich.-Désignation', 'Bem.-Remarques']]

    # Write results to Excel by adding or replacing the sheet "Änderungen" in the new file
    with pd.ExcelWriter(new_file, engine='openpyxl', mode='a') as writer:
      if 'Änderungen' in writer.book.sheetnames:
        del writer.book['Änderungen']

      df_merged.to_excel(writer, sheet_name='Änderungen', index=False)
      ws = writer.sheets['Änderungen']

      # Insert header row with comparison info
      ws.insert_rows(1)
      ws.cell(row=1, column=1, value='Verglichen:')
      ws.cell(row=1, column=2, value=f'{old_file_name} vs {new_file_name}')

      # Adjust column widths for selected columns
      auto_adjust_column_width(ws, ['A', 'D', 'E', 'F'])

      # Format header rows (row 1 and row 2) with neutral formatting
      for row in ws.iter_rows(min_row=1, max_row=2):
        for cell in row:
          cell.font = Font(bold=False, color='000000')
          cell.border = Border()

      # Format data rows (starting from row 3) based on change type
      for row in range(3, ws.max_row + 1):
        change_type = ws.cell(row=row, column=1).value
        if change_type == 'Hinzugefügt':
          font_color = 'FF0000'  # Red
        elif change_type == 'Geändert':
          font_color = '00B0F0'  # Blue
        elif change_type == 'Entfernt':
          font_color = 'FFC000'  # Yellow
        else:
          font_color = '000000'  # Black for Keine Änderung

        for col in range(1, ws.max_column + 1):
          cell = ws.cell(row=row, column=col)
          cell.font = Font(color=font_color, bold=False)
          cell.border = Border()

      writer.book.save(new_file)

    messagebox.showinfo("Erfolg", "Vergleich erfolgreich abgeschlossen und die Datei wurde gespeichert.")

  except Exception as e:
    messagebox.showerror("Fehler", f"Ein Fehler ist aufgetreten: {e}")

if __name__ == "__main__":
  main()
