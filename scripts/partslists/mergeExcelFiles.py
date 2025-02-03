# mergeExcelFiles.py
# Created by David Herren
# Version 2025-02-03

import pandas as pd
import os
from tkinter import filedialog, Tk
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

def auto_adjust_column_widths(ws, df):
  # Adjust column widths automatically
  ws.column_dimensions['A'].width = 30
  for col_idx, col_name in enumerate(df.columns, start=1):
    if col_idx == 1:
      continue
    max_length = max((len(str(cell)) for cell in df[col_name]), default=0)
    ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

def mergeExcelFiles():
  root = Tk()
  root.withdraw()

  # File selection
  file_paths = filedialog.askopenfilenames(
    title='Bitte wählen Sie die Excel-Dateien aus, die zusammengeführt werden sollen'
  )
  if not file_paths:
    print("Keine Dateien ausgewählt. Skript wird beendet.")
    return

  # Read and preprocess files
  all_data = []
  filenames = []
  expected_columns = ['Bezeich.-Désignation', 'Mat.', 'Bem.-Remarques', 'Stk', 'Art.No.']

  for file_path in file_paths:
    norm_path = os.path.normpath(file_path)
    df = pd.read_excel(norm_path)
    filenames.append(os.path.basename(norm_path))

    # Add missing columns at once
    for col in expected_columns:
      if col not in df.columns:
        df[col] = None

    if 'Pos.' in df.columns:
      df = df.drop(columns=['Pos.'])

    all_data.append(df)

  combined_df = pd.concat(all_data, ignore_index=True)

  # Create helper columns for grouping:
  # If Mat., Art.No. and Bem.-Remarques are all empty, we group by Bezeich.-Désignation only.
  # Otherwise, we group by Art.No., Bem.-Remarques and Bezeich.-Désignation so that different designations remain separate.
  mask = combined_df['Mat.'].isna() & combined_df['Art.No.'].isna() & combined_df['Bem.-Remarques'].isna()
  combined_df['grp_art_no'] = combined_df['Art.No.']
  combined_df['grp_bem'] = combined_df['Bem.-Remarques']
  combined_df['grp_designation'] = combined_df['Bezeich.-Désignation']
  combined_df.loc[mask, 'grp_art_no'] = None
  combined_df.loc[mask, 'grp_bem'] = None

  # Group data using the helper columns
  grouped = combined_df.groupby(
    ['grp_art_no', 'grp_bem', 'grp_designation'], dropna=False
  ).agg({
    'Stk': 'sum',
    'Mat.': 'first'
  }).reset_index()

  # Calculate group counts to set marker for merged rows
  count_df = combined_df.groupby(
    ['grp_art_no', 'grp_bem', 'grp_designation'], dropna=False
  ).size().rename('Count').reset_index()
  grouped = grouped.merge(
    count_df, on=['grp_art_no', 'grp_bem', 'grp_designation']
  )

  # Add marker column if more than one row was combined
  grouped[''] = grouped['Count'].apply(lambda x: '*' if x > 1 else '')
  grouped = grouped.drop(columns='Count')

  # Rename grouping columns to final column names
  grouped = grouped.rename(columns={
    'grp_art_no': 'Art.No.',
    'grp_bem': 'Bem.-Remarques',
    'grp_designation': 'Bezeich.-Désignation'
  })

  # Set final column order
  final_order = ['Art.No.', 'Stk', 'Bezeich.-Désignation', 'Mat.', 'Bem.-Remarques', '']
  grouped = grouped[final_order]

  # Ask for filename to save
  new_file_name = filedialog.asksaveasfilename(
    title='Bitte geben Sie den Namen für das neue Excel-File an',
    filetypes=[('Excel Files', '*.xlsx')]
  )
  if not new_file_name:
    print("Kein Dateiname eingegeben. Skript wird beendet.")
    return

  # Write to Excel
  wb = Workbook()
  ws = wb.active

  header_description = 'Zusammenfassung der Stk-Listen von ' + ', '.join(filenames)
  ws.append([header_description])

  for r in dataframe_to_rows(grouped, index=False, header=True):
    ws.append(r)

  # Adjust column widths
  auto_adjust_column_widths(ws, grouped)

  wb.save(new_file_name + '.xlsx')
  print(f"Datei wurde erfolgreich gespeichert: {new_file_name}.xlsx")

if __name__ == "__main__":
  mergeExcelFiles()
